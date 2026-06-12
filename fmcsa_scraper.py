"""
FMCSA Carrier Data Extractor
============================
Queries the official US DOT public data portal (data.transportation.gov)
for motor carrier census data and filters for owner-operators
(small fleet operators) by state.

Dataset: SMS Input - Motor Carrier Census Information
  https://data.transportation.gov/Public-Safety/SMS-Input-Motor-Carrier-Census-Information/kjg3-diqy
  Published by the Federal Motor Carrier Safety Administration (FMCSA) as open government data.
"""

import csv
import io
import os
import sys
from dataclasses import dataclass, fields
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# -- Configuration ---------------------------------------------------------

SODA_API = "https://data.transportation.gov/resource/kjg3-diqy.csv"
DEFAULT_OUTPUT = "truck_owners_usa.xlsx"
MAX_POWER_UNITS = 2

# SODA query limit per page
SODA_PAGE_SIZE = 50_000
MAX_RECORDS = 0  # 0 = unlimited

# -- Data model -------------------------------------------------------------


@dataclass
class CarrierRecord:
    dot_number: str = ""
    legal_name: str = ""
    dba_name: str = ""
    carrier_operation: str = ""
    hm_flag: str = ""
    pc_flag: str = ""
    phy_street: str = ""
    phy_city: str = ""
    phy_state: str = ""
    phy_zip: str = ""
    phy_country: str = ""
    mailing_street: str = ""
    mailing_city: str = ""
    mailing_state: str = ""
    mailing_zip: str = ""
    mailing_country: str = ""
    telephone: str = ""
    fax: str = ""
    email_address: str = ""
    mcs150_date: str = ""
    nbr_power_unit: str = ""
    driver_total: str = ""
    authorized_for_hire: str = ""
    private_only: str = ""
    exempt_for_hire: str = ""
    op_other: str = ""


DISPLAY_NAMES = [
    "USDOT #",
    "Legal Name",
    "DBA Name",
    "Operation",
    "HM Flag",
    "PC Flag",
    "Street",
    "City",
    "State",
    "ZIP",
    "Country",
    "Mailing Street",
    "Mailing City",
    "Mailing State",
    "Mailing ZIP",
    "Mailing Country",
    "Phone",
    "Fax",
    "Email",
    "MCS-150 Date",
    "Power Units",
    "Drivers",
    "Authorized for Hire",
    "Private Only",
    "Exempt for Hire",
    "Other Ops",
]


# -- API helpers ------------------------------------------------------------


def build_soql(
    states: list[str] | None = None,
    offset: int = 0,
    limit: int = SODA_PAGE_SIZE,
) -> str:
    conditions = ["carrier_operation = 'A'"]
    if states:
        quoted = ",".join(f"'{s.upper()}'" for s in states)
        conditions.append(f"phy_state in({quoted})")
    fields_list = ",".join(f.name for f in fields(CarrierRecord))
    soql = (
        f"SELECT {fields_list} WHERE {' AND '.join(conditions)}"
        f" ORDER BY dot_number"
        f" OFFSET {offset} LIMIT {limit}"
    )
    return soql


def fetch_page(soql: str) -> list[CarrierRecord]:
    resp = requests.get(
        SODA_API,
        params={"$query": soql},
        headers={"Accept": "text/csv"},
        timeout=120,
    )
    if resp.status_code == 404:
        return []
    resp.raise_for_status()
    text = resp.text
    if not text.strip():
        return []
    reader = csv.DictReader(io.StringIO(text))
    records: list[CarrierRecord] = []
    valid_fields = {f.name for f in fields(CarrierRecord)}
    for row in reader:
        mapped = {
            k.strip(): (v or "").strip()
            for k, v in row.items()
            if k.strip() in valid_fields
        }
        records.append(CarrierRecord(**mapped))
    return records


def fetch_all(
    states: list[str] | None,
    max_records: int = MAX_RECORDS,
) -> list[CarrierRecord]:
    all_records: list[CarrierRecord] = []
    offset = 0
    page = 0
    while True:
        soql = build_soql(states, offset, SODA_PAGE_SIZE)
        print(f"  Page {page + 1}: offset={offset:,} limit={SODA_PAGE_SIZE:,}")
        page_records = fetch_page(soql)
        if not page_records:
            break
        all_records.extend(page_records)
        offset += len(page_records)
        page += 1
        if max_records and len(all_records) >= max_records:
            all_records = all_records[:max_records]
            break
        if len(page_records) < SODA_PAGE_SIZE:
            break
    return all_records


# -- Filtering --------------------------------------------------------------


def is_owner_operator(rec: CarrierRecord, max_power_units: int = MAX_POWER_UNITS) -> bool:
    try:
        pu = int(rec.nbr_power_unit)
    except (ValueError, TypeError):
        return False
    if pu < 1 or pu > max_power_units:
        return False
    # Prefer authorized for hire
    if rec.authorized_for_hire != "true":
        return False
    return True


# -- Excel export -----------------------------------------------------------


def write_xlsx(
    records: list[CarrierRecord],
    path: str | Path,
    state_filters: list[str] | None,
    max_power_units: int = MAX_POWER_UNITS,
):
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Data sheet
    ws = wb.active
    ws.title = "Carriers"
    field_names = [f.name for f in fields(CarrierRecord)]

    for col_idx, name in enumerate(DISPLAY_NAMES, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.freeze_panes = "A2"

    for row_idx, rec in enumerate(records, 2):
        for col_idx, fname in enumerate(field_names, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=getattr(rec, fname, ""))
            cell.border = thin_border

    for col_idx, name in enumerate(DISPLAY_NAMES, 1):
        max_len = len(name)
        for row_idx in range(2, min(len(records) + 2, 200)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 50))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Field").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="Value").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill

    summary_data = [
        ("Extraction Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Data Source", "data.transportation.gov - SMS Input - Motor Carrier Census Information"),
        ("Total Records", str(len(records))),
        ("Filter", f"Interstate (A), Authorized for Hire, Power units 1-{max_power_units}"),
    ]
    if state_filters:
        summary_data.append(("State Filter(s)", ", ".join(state_filters)))
    for i, (k, v) in enumerate(summary_data, 2):
        ws2.cell(row=i, column=1, value=k).border = thin_border
        ws2.cell(row=i, column=2, value=v).border = thin_border
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 80

    wb.save(path)
    print(f"  Workbook saved: {path}")


# -- CLI --------------------------------------------------------------------


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="FMCSA Truck Owner Data Extractor - public US carrier census via data.transportation.gov",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python fmcsa_scraper.py\n"
            "  python fmcsa_scraper.py --states CA TX FL\n"
            "  python fmcsa_scraper.py --states NY --output ny_truck_owners.xlsx\n"
            "  python fmcsa_scraper.py --states TX --max-records 100\n"
        ),
    )
    parser.add_argument(
        "--states",
        nargs="+",
        default=None,
        help="Two-letter state code(s) (e.g. CA TX FL). Omit for all 50 states.",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT,
        help=f"Output .xlsx (default: {DEFAULT_OUTPUT})",
    )
    parser.add_argument(
        "--max-power-units",
        type=int,
        default=MAX_POWER_UNITS,
        help=f"Max power units to qualify as owner-operator (default: {MAX_POWER_UNITS})",
    )
    parser.add_argument(
        "--max-records",
        type=int,
        default=0,
        help="Max records to fetch (0 = unlimited). Use for testing.",
    )
    args = parser.parse_args()

    max_power_units = args.max_power_units
    max_records = args.max_records or 0

    print("=" * 60)
    print("  FMCSA Truck Owner Data Extractor")
    print("  Source: data.transportation.gov - US DOT open data portal")
    print("  Dataset: SMS Input - Motor Carrier Census Information")
    print("=" * 60)

    print(f"\n[1/3] Querying API ...")
    print(f"  Server filter: Interstate carriers (operation=A)")
    if args.states:
        print(f"  State filter: {', '.join(s.upper() for s in args.states)}")
    all_records = fetch_all(states=args.states, max_records=max_records)
    print(f"  Downloaded: {len(all_records):,} raw records")

    if not all_records:
        print("  No records returned. Check network or try again.")
        sys.exit(1)

    print(f"\n[2/3] Filtering for owner-operators (<={max_power_units} power unit(s), authorized for hire) ...")
    records = [r for r in all_records if is_owner_operator(r, max_power_units)]
    print(f"  Owner-operators found: {len(records):,}")

    if not records:
        print("\n  No matching records. Try expanding --max-power-units.")
        sys.exit(0)

    print(f"\n[3/3] Writing workbook ...")
    write_xlsx(records, args.output, args.states, max_power_units)

    print(f"\n  Done! {len(records):,} truck owners saved to: {args.output}")
    print("=" * 60)


if __name__ == "__main__":
    main()
